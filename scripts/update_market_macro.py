from __future__ import annotations

import csv
import html
import json
import re
import tempfile
from datetime import datetime, timedelta, timezone
from xml.etree import ElementTree as ET
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen


START_DATE = "1965-01-01"
GDP_START_DATE = "1981-01-01"
LIQUIDITY_START_DATE = "1981-01-01"
FX_START_DATE = "1971-01-01"
FOOD_START_DATE = "2001-01-01"
FRED_GRAPH_BASE = "https://fred.stlouisfed.org/graph/fredgraph.csv?id="
FRED_GATEWAY_BASE = "https://www.ivo-welch.info/cgi-bin/fredwrap?symbol="
STREETSTATS_BASE = "https://streetstats.finance"
FISCALDATA_TGA_URL = (
    "https://api.fiscaldata.treasury.gov/services/api/fiscal_service/v1/accounting/dts/"
    "operating_cash_balance"
)
WORLD_BANK_PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/"
    "related/CMO-Historical-Data-Monthly.xlsx"
)
RANGES = [
    {"key": "1m", "label": "1M"},
    {"key": "3m", "label": "3M"},
    {"key": "6m", "label": "6M"},
    {"key": "ytd", "label": "YTD"},
    {"key": "1y", "label": "1Y"},
    {"key": "3y", "label": "3Y"},
    {"key": "5y", "label": "5Y"},
    {"key": "max", "label": "Max"},
]

US_TREASURY_SERIES = {
    "us2y": ("US 2Y", "#111827"),
    "us5y": ("US 5Y", "#0f766e"),
    "us10y": ("US 10Y", "#2563eb"),
    "us30y": ("US 30Y", "#dc2626"),
}

YAHOO_SERIES = {
    "dxy": ("DX-Y.NYB", "Dollar Index", "#111827"),
    "wti": ("CL=F", "WTI", "#7c3aed"),
    "brent": ("BZ=F", "Brent", "#2563eb"),
    "gold": ("GC=F", "Gold", "#d97706"),
    "silver": ("SI=F", "Silver", "#6b7280"),
    "copper": ("HG=F", "Copper", "#b45309"),
    "iron_ore": ("TIO=F", "Iron Ore 62% Fe", "#b45309"),
    "lng_jkm": ("JKM=F", "JKM LNG", "#2563eb"),
    "corn": ("ZC=F", "Corn", "#f59e0b"),
    "soybeans": ("ZS=F", "Soybeans", "#16a34a"),
    "wheat_hrw": ("KE=F", "Wheat HRW", "#dc2626"),
    "wheat_srw": ("ZW=F", "Wheat SRW", "#7c3aed"),
    "sugar": ("SB=F", "Raw Sugar", "#0f766e"),
}


def fetch_text(url: str) -> str:
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    last_error: Exception | None = None
    for _ in range(3):
        try:
            with urlopen(request, timeout=60) as response:  # nosec B310 - fixed public endpoints
                return response.read().decode("utf-8-sig")
        except Exception as error:  # pragma: no cover - network variability
            last_error = error
    raise RuntimeError(f"Failed to fetch {url}") from last_error


def fetch_text_with_headers(url: str, headers: dict[str, str]) -> str:
    request = Request(url, headers=headers)
    last_error: Exception | None = None
    for _ in range(3):
        try:
            with urlopen(request, timeout=60) as response:  # nosec B310 - fixed public endpoints
                return response.read().decode("utf-8-sig")
        except Exception as error:  # pragma: no cover - network variability
            last_error = error
    raise RuntimeError(f"Failed to fetch {url}") from last_error


def fetch_json(url: str) -> dict:
    return json.loads(fetch_text(url))


def read_existing_payload() -> dict:
    output_path = Path(__file__).resolve().parents[1] / "data" / "market-macro-data.js"
    if not output_path.exists():
        return {}
    raw = output_path.read_text(encoding="utf-8").strip()
    prefix = "window.marketMacroData = "
    if raw.startswith(prefix):
        raw = raw[len(prefix) :]
    return json.loads(raw.rstrip(";"))


def existing_macro_series(panel_key: str, series_key: str) -> tuple[list[str], list[float]]:
    try:
        item = read_existing_payload()["panels"][panel_key]["series"][series_key]
        dates = item.get("dates") or []
        values = item.get("values") or []
        if dates and values:
            return list(dates), [float(value) for value in values]
    except Exception:
        pass
    return [], []


def parse_fred_series_or_existing(
    series_id: str,
    start_date: str,
    panel_key: str,
    series_key: str,
) -> tuple[list[str], list[float]]:
    try:
        return parse_fred_series(series_id, start_date)
    except Exception as error:  # pragma: no cover - network variability
        dates, values = existing_macro_series(panel_key, series_key)
        if dates:
            print(f"Using existing {panel_key}.{series_key} after {series_id} fetch failed: {error}")
            return dates, values
        raise


def fred_csv_url(series_id: str) -> str:
    return f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"


def fred_csv_urls(series_id: str, start_date: str = START_DATE) -> list[str]:
    return [
        f"{FRED_GATEWAY_BASE}{series_id}",
        f"{FRED_GRAPH_BASE}{series_id}&cosd={start_date}",
    ]


def yahoo_chart_url(symbol: str) -> str:
    encoded = quote(symbol, safe="")
    period1 = int(datetime.strptime(START_DATE, "%Y-%m-%d").replace(tzinfo=timezone.utc).timestamp())
    period2 = int(datetime.now(timezone.utc).timestamp())
    return (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{encoded}"
        f"?period1={period1}&period2={period2}&interval=1d&includeAdjustedClose=true&events=div%2Csplits"
    )


def parse_us_treasury_yields() -> dict[str, tuple[list[str], list[float]]]:
    series = {key: ([], []) for key in ("us2y", "us5y", "us10y", "us30y")}
    ns = {
        "atom": "http://www.w3.org/2005/Atom",
        "d": "http://schemas.microsoft.com/ado/2007/08/dataservices",
        "m": "http://schemas.microsoft.com/ado/2007/08/dataservices/metadata",
    }
    start_year = int(START_DATE[:4])
    current_year = datetime.now(timezone.utc).year
    for year in range(start_year, current_year + 1):
        url = (
            "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/pages/xml"
            f"?data=daily_treasury_yield_curve&field_tdr_date_value={year}"
        )
        root = ET.fromstring(fetch_text(url))
        for entry in root.findall("atom:entry", ns):
            properties = entry.find("atom:content/m:properties", ns)
            if properties is None:
                continue
            raw_date = properties.findtext("d:NEW_DATE", default="", namespaces=ns)
            if not raw_date:
                continue
            date_key = raw_date.split("T", 1)[0]
            if date_key < START_DATE:
                continue
            for key, field_name in {
                "us2y": "BC_2YEAR",
                "us5y": "BC_5YEAR",
                "us10y": "BC_10YEAR",
                "us30y": "BC_30YEAR",
            }.items():
                raw_value = properties.findtext(f"d:{field_name}", default="", namespaces=ns).strip()
                if raw_value in {"", "N/A"}:
                    continue
                series[key][0].append(date_key)
                series[key][1].append(round(float(raw_value), 4))
    return series


def parse_japan_yields() -> dict[str, tuple[list[str], list[float]]]:
    series_map: dict[str, dict[str, float]] = {key: {} for key in ("jp2y", "jp10y", "jp30y")}

    mof_url = "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/historical/jgbcme_all.csv"
    mof_lines = fetch_text(mof_url).splitlines()
    header_index = next(index for index, line in enumerate(mof_lines) if line.startswith("Date,"))
    mof_reader = csv.DictReader(mof_lines[header_index:])
    for row in mof_reader:
        raw_date = (row.get("Date") or "").strip()
        if not raw_date:
            continue
        try:
            date_key = datetime.strptime(raw_date, "%Y/%m/%d").strftime("%Y-%m-%d")
        except ValueError:
            continue
        if date_key < START_DATE:
            continue
        for key, column in {"jp2y": "2Y", "jp10y": "10Y", "jp30y": "30Y"}.items():
            raw_value = (row.get(column) or "").strip()
            if raw_value in {"", "-", "--"}:
                continue
            series_map[key][date_key] = round(float(raw_value), 4)

    jbts_url = "https://www.bb.jbts.co.jp/en/historical/main_rate.html"
    text = fetch_text(jbts_url)
    table_blocks = re.findall(r'<table class="tbCore">(.*?)</table>', text, flags=re.S | re.I)
    for block in table_blocks:
        row_matches = re.findall(r"<tr>(.*?)</tr>", block, flags=re.S | re.I)
        for row_html in row_matches:
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row_html, flags=re.S | re.I)
            if len(cells) < 7:
                continue
            values = [
                html.unescape(re.sub(r"<[^>]+>", "", cell).strip())
                for cell in cells
            ]
            raw_date = values[0]
            if raw_date.lower() == "date" or not raw_date:
                continue
            try:
                date_key = datetime.strptime(raw_date, "%Y/%m/%d").strftime("%Y-%m-%d")
            except ValueError:
                continue
            if date_key < START_DATE:
                continue
            for key, index in {"jp30y": 2, "jp10y": 4, "jp2y": 6}.items():
                raw_value = values[index]
                if raw_value in {"", "-", "--"}:
                    continue
                series_map[key][date_key] = round(float(raw_value), 4)

    series = {}
    for key, by_date in series_map.items():
        dates = sorted(by_date)
        series[key] = (dates, [by_date[date] for date in dates])
    return series


def parse_yahoo_series(symbol: str) -> tuple[list[str], list[float]]:
    payload = fetch_json(yahoo_chart_url(symbol))
    result = payload["chart"]["result"][0]
    timestamps = result.get("timestamp") or []
    quote_data = (result.get("indicators") or {}).get("quote") or [{}]
    adjclose_data = (result.get("indicators") or {}).get("adjclose") or [{}]
    closes = adjclose_data[0].get("adjclose") or quote_data[0].get("close") or []

    dates: list[str] = []
    values: list[float] = []
    for timestamp, close in zip(timestamps, closes):
        if close is None:
            continue
        date_key = (datetime(1970, 1, 1, tzinfo=timezone.utc) + timedelta(seconds=timestamp)).strftime("%Y-%m-%d")
        if date_key < START_DATE:
            continue
        dates.append(date_key)
        values.append(round(float(close), 4))
    return dates, values


def parse_world_bank_monthly_prices() -> dict[str, tuple[list[str], list[float]]]:
    import openpyxl

    workbook_path = Path(tempfile.gettempdir()) / "CMO-Historical-Data-Monthly.xlsx"
    workbook_path.write_bytes(
        urlopen(Request(WORLD_BANK_PINK_SHEET_URL, headers={"User-Agent": "Mozilla/5.0"}), timeout=60).read()
    )
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook["Monthly Prices"]
    code_row = next(worksheet.iter_rows(min_row=7, max_row=7, values_only=True))
    target_columns = {
        "brent": "CRUDE_BRENT",
        "dubai": "CRUDE_DUBAI",
        "wti": "CRUDE_WTI",
        "gold": "GOLD",
        "silver": "SILVER",
        "copper": "COPPER",
        "iron_ore": "IRON_ORE",
        "nickel": "NICKEL",
        "zinc": "Zinc",
    }
    column_index = {
        key: code_row.index(code)
        for key, code in target_columns.items()
        if code in code_row
    }
    series = {key: ([], []) for key in target_columns}
    for row in worksheet.iter_rows(min_row=8, values_only=True):
        raw_period = row[0]
        if not raw_period:
            continue
        period = str(raw_period)
        match = re.match(r"^(\d{4})M(\d{2})$", period)
        if not match:
            continue
        date_key = f"{match.group(1)}-{match.group(2)}-01"
        if date_key < START_DATE:
            continue
        for key, index in column_index.items():
            value = row[index]
            if not isinstance(value, (int, float)):
                continue
            series[key][0].append(date_key)
            series[key][1].append(round(float(value), 4))
    return series


def parse_fred_series(series_id: str, start_date: str = START_DATE) -> tuple[list[str], list[float]]:
    last_error: Exception | None = None
    for url in fred_csv_urls(series_id, start_date):
        try:
            reader = csv.DictReader(fetch_text(url).splitlines())
            dates: list[str] = []
            values: list[float] = []
            for row in reader:
                lower_row = {str(key).lower(): value for key, value in row.items()}
                raw_date = (row.get("DATE") or lower_row.get("yyyymmdd") or "").strip()
                raw_value = (row.get(series_id) or lower_row.get(series_id.lower()) or "").strip()
                if not raw_date or raw_value in {"", "."}:
                    continue
                if len(raw_date) == 8 and raw_date.isdigit():
                    date_key = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:8]}"
                else:
                    date_key = raw_date[:10]
                if date_key < start_date:
                    continue
                dates.append(date_key)
                values.append(round(float(raw_value), 4))
            if dates:
                return dates, values
        except Exception as error:  # pragma: no cover - network variability
            last_error = error
    if last_error:
        raise last_error
    return [], []


def parse_real_gdp_annualized_series() -> tuple[list[str], list[float]]:
    export_url = "https://govspending.org/api/export/fred/A191RL1Q225SBEA.csv"
    try:
        rows = [
            line
            for line in fetch_text(export_url).splitlines()
            if line and not line.startswith("#")
        ]
        reader = csv.DictReader(rows)
        dates: list[str] = []
        values: list[float] = []
        for row in reader:
            date_key = (row.get("date") or row.get("DATE") or "")[:10]
            raw_value = (row.get("A191RL1Q225SBEA") or "").strip()
            if date_key < GDP_START_DATE or raw_value in {"", "."}:
                continue
            dates.append(date_key)
            values.append(round(float(raw_value), 4))
        if len(dates) >= 180:
            return dates, values
    except Exception:
        pass
    return parse_fred_series("A191RL1Q225SBEA", GDP_START_DATE)


def build_daily_forward_series(
    source_dates: list[str],
    source_values: list[float],
    start_date: str,
    end_date: str | None = None,
) -> tuple[list[str], list[float | None]]:
    if end_date is None:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    cursor = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    dates: list[str] = []
    values: list[float | None] = []
    source_index = 0
    carried_value: float | None = None
    while cursor <= end:
        date_key = cursor.strftime("%Y-%m-%d")
        while source_index < len(source_dates) and source_dates[source_index] <= date_key:
            carried_value = source_values[source_index]
            source_index += 1
        dates.append(date_key)
        values.append(carried_value)
        cursor += timedelta(days=1)
    return dates, values


def parse_streetstats_global_m2_proxy() -> tuple[list[str], list[float | None]]:
    token_payload = json.loads(fetch_text(f"{STREETSTATS_BASE}/api/token"))
    token = token_payload["token"]
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": f"{STREETSTATS_BASE}/liquidity/money/",
        "Authorization": f"Bearer {token}",
    }
    url = (
        f"{STREETSTATS_BASE}/api/econ/columns"
        "?id=Date&id=Global_M2&id=US_M2&id=Euro_M2&id=China_M2&id=Japan_M2"
    )
    rows = json.loads(fetch_text_with_headers(url, headers))
    source_dates: list[str] = []
    source_values: list[float] = []
    for row in rows:
        raw_date = str(row.get("Date") or "").strip()
        raw_value = row.get("Global_M2")
        if not raw_date or raw_value in {"", None}:
            continue
        try:
            date_key = datetime.strptime(raw_date, "%m/%d/%Y").strftime("%Y-%m-%d")
        except ValueError:
            continue
        source_dates.append(date_key)
        source_values.append(round(float(raw_value), 2))
    return build_daily_forward_series(source_dates, source_values, LIQUIDITY_START_DATE)


def parse_sofr_iorb_spread_bps() -> tuple[list[str], list[float | None]]:
    sofr_dates, sofr_values = parse_fred_series("SOFR", LIQUIDITY_START_DATE)
    iorb_dates, iorb_values = parse_fred_series("IORB", LIQUIDITY_START_DATE)
    try:
        ioer_dates, ioer_values = parse_fred_series("IOER", LIQUIDITY_START_DATE)
        iorb_dates, iorb_values = merge_series_prefer_recent(
            (iorb_dates, iorb_values),
            (ioer_dates, ioer_values),
        )
    except Exception:
        pass
    spread_dates, spread_values = build_spread_series(
        sofr_dates,
        sofr_values,
        iorb_dates,
        iorb_values,
    )
    spread_bps = [round(value * 100, 2) for value in spread_values]
    return build_daily_forward_series(spread_dates, spread_bps, LIQUIDITY_START_DATE)


def parse_tga_daily_balance() -> tuple[list[str], list[float]]:
    base_url = (
        f"{FISCALDATA_TGA_URL}"
        f"?filter=record_date:gte:{LIQUIDITY_START_DATE}"
        "&fields=record_date,account_type,open_today_bal"
        "&sort=record_date"
        "&page[size]=10000"
    )
    accepted_types = {
        "Treasury General Account (TGA)",
        "Treasury General Account (TGA) Closing Balance",
    }
    rows: list[dict[str, str]] = []
    page = 1
    while True:
        payload = fetch_json(f"{base_url}&page[number]={page}")
        rows.extend(payload.get("data") or [])
        total_pages = int((payload.get("meta") or {}).get("total-pages") or page)
        if page >= total_pages:
            break
        page += 1

    values_by_date: dict[str, float] = {}
    for row in rows:
        account_type = str(row.get("account_type") or "").strip()
        raw_date = str(row.get("record_date") or "")[:10]
        raw_value = str(row.get("open_today_bal") or "").strip()
        if account_type not in accepted_types or not raw_date or raw_value in {"", "null", "."}:
            continue
        values_by_date[raw_date] = round(float(raw_value) / 1000, 2)

    dates = sorted(values_by_date)
    return dates, [values_by_date[date] for date in dates]


def build_spread_series(
    first_dates: list[str],
    first_values: list[float],
    second_dates: list[str],
    second_values: list[float],
) -> tuple[list[str], list[float]]:
    first_by_date = dict(zip(first_dates, first_values))
    second_by_date = dict(zip(second_dates, second_values))
    dates = sorted(set(first_by_date) & set(second_by_date))
    values = [round(first_by_date[date] - second_by_date[date], 4) for date in dates]
    return dates, values


def build_policy_rate_series() -> tuple[list[str], list[float]]:
    fedfunds_dates, fedfunds_values = parse_fred_series("FEDFUNDS")
    target_dates, target_values = parse_fred_series("DFEDTAR")
    target_upper_dates, target_upper_values = parse_fred_series("DFEDTARU")
    return merge_series_prefer_recent(
        (target_upper_dates, target_upper_values),
        merge_series_prefer_recent(
            (target_dates, target_values),
            (fedfunds_dates, fedfunds_values),
        ),
    )


def build_series_item(
    label: str,
    color: str,
    dates: list[str],
    values: list[float],
    dash: list[int] | None = None,
) -> dict[str, object]:
    item: dict[str, object] = {
        "name": label,
        "color": color,
        "dates": dates,
        "values": values,
    }
    if dash:
        item["dash"] = dash
    return item


def filter_series_start(
    dates: list[str],
    values: list[float],
    start_date: str,
) -> tuple[list[str], list[float]]:
    filtered = [(date, value) for date, value in zip(dates, values) if date >= start_date]
    return [date for date, _ in filtered], [value for _, value in filtered]


def merge_series_prefer_recent(
    primary: tuple[list[str], list[float]],
    fallback: tuple[list[str], list[float]],
) -> tuple[list[str], list[float]]:
    merged: dict[str, float] = {}
    for date, value in zip(fallback[0], fallback[1]):
        merged[date] = value
    for date, value in zip(primary[0], primary[1]):
        merged[date] = value
    dates = sorted(merged)
    return dates, [merged[date] for date in dates]


def main() -> None:
    us_series = parse_us_treasury_yields()
    japan_series = parse_japan_yields()
    rates_series = {
        key: build_series_item(label, color, *us_series[key])
        for key, (label, color) in US_TREASURY_SERIES.items()
    }
    rates_series["jp2y"] = build_series_item("Japan 2Y", "#111827", *japan_series["jp2y"], [6, 4])
    rates_series["jp10y"] = build_series_item("Japan 10Y", "#2563eb", *japan_series["jp10y"], [6, 4])
    rates_series["jp30y"] = build_series_item("Japan 30Y", "#dc2626", *japan_series["jp30y"], [6, 4])
    fed_funds_dates, fed_funds_values = build_policy_rate_series()
    gdp_dates, gdp_values = parse_real_gdp_annualized_series()
    liquidity_fed_dates, liquidity_fed_values = build_daily_forward_series(
        *filter_series_start(fed_funds_dates, fed_funds_values, LIQUIDITY_START_DATE),
        LIQUIDITY_START_DATE,
    )
    dgs2_dates, dgs2_values = parse_fred_series("DGS2", LIQUIDITY_START_DATE)
    liquidity_us2y_dates, liquidity_us2y_values = build_daily_forward_series(
        dgs2_dates,
        dgs2_values,
        LIQUIDITY_START_DATE,
    )
    global_m2_dates, global_m2_values = parse_streetstats_global_m2_proxy()
    tga_dates, tga_values = parse_tga_daily_balance()
    sofr_iorb_dates, sofr_iorb_values = parse_sofr_iorb_spread_bps()
    inflation_5y_dates, inflation_5y_values = parse_fred_series("T5YIE")
    real_5y_dates, real_5y_values = build_spread_series(
        us_series["us5y"][0],
        us_series["us5y"][1],
        inflation_5y_dates,
        inflation_5y_values,
    )
    if not real_5y_dates:
        real_5y_dates, real_5y_values = parse_fred_series("DFII5")

    dxy_dates, dxy_values = parse_yahoo_series("DX-Y.NYB")
    dxy_dates, dxy_values = filter_series_start(dxy_dates, dxy_values, FX_START_DATE)
    jpy_usd_dates, jpy_usd_values = parse_fred_series("DEXJPUS", FX_START_DATE)
    krw_usd_dates, krw_usd_values = parse_fred_series("DEXKOUS", FX_START_DATE)
    chf_usd_dates, chf_usd_values = parse_fred_series("DEXSZUS", FX_START_DATE)
    cny_usd_dates, cny_usd_values = parse_fred_series("DEXCHUS", FX_START_DATE)
    gbp_usd_dates, gbp_usd_values = parse_fred_series("DEXUSUK", FX_START_DATE)
    eur_usd_dates, eur_usd_values = parse_fred_series("DEXUSEU", FX_START_DATE)
    long_commodity_series = parse_world_bank_monthly_prices()
    wti_dates, wti_values = parse_yahoo_series("CL=F")
    brent_dates, brent_values = parse_yahoo_series("BZ=F")
    dubai_dates, dubai_values = long_commodity_series["dubai"]
    henry_hub_dates, henry_hub_values = parse_fred_series_or_existing("DHHNGSP", START_DATE, "natural_gas", "henry_hub")
    gold_dates, gold_values = parse_yahoo_series("GC=F")
    silver_dates, silver_values = parse_yahoo_series("SI=F")
    copper_dates, copper_values = parse_yahoo_series("HG=F")
    uranium_dates, uranium_values = parse_fred_series_or_existing("PURANUSDM", START_DATE, "strategic", "uranium")
    iron_ore_dates, iron_ore_values = parse_yahoo_series("TIO=F")
    nickel_dates, nickel_values = long_commodity_series["nickel"]
    zinc_dates, zinc_values = long_commodity_series["zinc"]
    lng_jkm_dates, lng_jkm_values = parse_yahoo_series("JKM=F")
    corn_dates, corn_values = parse_yahoo_series("ZC=F")
    soybeans_dates, soybeans_values = parse_yahoo_series("ZS=F")
    wheat_hrw_dates, wheat_hrw_values = parse_yahoo_series("KE=F")
    wheat_srw_dates, wheat_srw_values = parse_yahoo_series("ZW=F")
    sugar_dates, sugar_values = parse_yahoo_series("SB=F")
    corn_dates, corn_values = filter_series_start(corn_dates, corn_values, FOOD_START_DATE)
    soybeans_dates, soybeans_values = filter_series_start(soybeans_dates, soybeans_values, FOOD_START_DATE)
    wheat_hrw_dates, wheat_hrw_values = filter_series_start(wheat_hrw_dates, wheat_hrw_values, FOOD_START_DATE)
    wheat_srw_dates, wheat_srw_values = filter_series_start(wheat_srw_dates, wheat_srw_values, FOOD_START_DATE)
    sugar_dates, sugar_values = filter_series_start(sugar_dates, sugar_values, FOOD_START_DATE)
    wti_dates, wti_values = merge_series_prefer_recent(
        (wti_dates, wti_values),
        long_commodity_series["wti"],
    )

    panels = {
        "policy": {
            "title": "Policy Rate / Real Rate",
            "subtitle": "Fed policy-rate proxy uses monthly effective Fed Funds before 1982, the Fed Funds Target Rate from 1982 to 2008, and the target-range upper bound after 2008. 5Y breakeven inflation starts in 2003.",
            "source": "FRED / US Treasury",
            "mode": "raw",
            "yAxisLabel": "%",
            "formatter": "percent2",
            "series": {
                "fed_funds": build_series_item("Fed Policy Rate", "#111827", fed_funds_dates, fed_funds_values),
                "inflation_5y": build_series_item("5Y Inflation Expectation", "#f97316", inflation_5y_dates, inflation_5y_values),
                "real_5y": build_series_item("Real 5Y (5Y - 5Y Inflation Exp.)", "#dc2626", real_5y_dates, real_5y_values),
            },
        },
        "gdp": {
            "title": "US GDP",
            "subtitle": "Quarterly real GDP percent change from preceding period, seasonally adjusted annual rate, from 1981 Q1.",
            "source": "FRED A191RL1Q225SBEA / BEA",
            "mode": "raw",
            "connectGaps": True,
            "fillMissing": "forward",
            "yAxisLabel": "Annualized QoQ %",
            "formatter": "percent2",
            "series": {
                "real_gdp_annualized": {
                    **build_series_item("Real GDP QoQ SAAR", "#8b5cf6", gdp_dates, gdp_values),
                    "fillForward": True,
                },
            },
        },
        "rates": {
            "title": "US & Japan Sovereign Yields",
            "subtitle": "Daily 2Y / 10Y / 30Y. US Treasury from Treasury XML, Japan JGB from JBTS.",
            "source": "US Treasury / Japan Bond Trading",
            "mode": "raw",
            "yAxisLabel": "Yield %",
            "formatter": "percent2",
            "series": rates_series,
        },
        "liquidity_global_m2": {
            "title": "Global Money Supply Proxy",
            "subtitle": "Daily forward-filled proxy for GLMOSUPP-style global M2 in USD, using US, Eurozone, China, and Japan M2 converted to USD where available.",
            "source": "StreetStats public Global M2 proxy / central bank sources",
            "mode": "raw",
            "connectGaps": False,
            "yAxisLabel": "USD bn",
            "formatter": "number1",
            "series": {
                "global_m2_usd": build_series_item("Global M2 Proxy (USD)", "#2563eb", global_m2_dates, global_m2_values),
            },
        },
        "liquidity_sofr_iorb": {
            "title": "SOFR - IORB",
            "subtitle": "Daily SOFR less the Fed reserve-balance administered rate, shown in basis points. IOER is used as the historical reserve-rate bridge before IORB.",
            "source": "FRED SOFR / IORB / IOER",
            "mode": "raw",
            "connectGaps": False,
            "yAxisLabel": "bps",
            "formatter": "number1",
            "series": {
                "sofr_iorb": build_series_item("SOFR - IORB", "#dc2626", sofr_iorb_dates, sofr_iorb_values),
            },
        },
        "liquidity_tga": {
            "title": "Treasury General Account",
            "subtitle": "Daily TGA cash balance from the Daily Treasury Statement, shown in USD billions.",
            "source": "U.S. Treasury FiscalData Daily Treasury Statement",
            "mode": "raw",
            "connectGaps": False,
            "yAxisLabel": "USD bn",
            "formatter": "number1",
            "series": {
                "tga_balance": build_series_item("TGA Balance", "#0f766e", tga_dates, tga_values),
            },
        },
        "liquidity_policy_2y": {
            "title": "Fed Policy vs US 2Y",
            "subtitle": "Step-like Fed policy-rate proxy and US 2Y Treasury yield, forward-filled across non-trading days for continuous comparison from 1981-01-01.",
            "source": "FRED DFEDTARU / DFEDTAR / FEDFUNDS / DGS2",
            "mode": "raw",
            "connectGaps": False,
            "yAxisLabel": "%",
            "formatter": "percent2",
            "series": {
                "fed_policy_rate": build_series_item("Fed Policy Rate", "#111827", liquidity_fed_dates, liquidity_fed_values),
                "us2y": build_series_item("US 2Y Treasury", "#2563eb", liquidity_us2y_dates, liquidity_us2y_values),
            },
        },
        "dxy": {
            "title": "Dollar Index",
            "subtitle": "Daily DXY close from Yahoo Finance.",
            "source": "Yahoo Finance",
            "mode": "raw",
            "yAxisLabel": "Index",
            "formatter": "number1",
            "series": {
                "dxy": build_series_item("Dollar Index", "#111827", dxy_dates, dxy_values),
            },
        },
        "fx_dashboard": {
            "title": "FX Dashboard",
            "subtitle": "Dollar Index plus major USD exchange-rate series from 1971-01-01 where available. Each line is normalized to 100 at the selected start date.",
            "source": "FRED / Yahoo Finance",
            "mode": "normalized",
            "yAxisLabel": "Start = 100",
            "formatter": "number1",
            "series": {
                "dxy": build_series_item("Dollar Index", "#111827", dxy_dates, dxy_values),
                "jpy_usd": build_series_item("JPY/USD", "#dc2626", jpy_usd_dates, jpy_usd_values),
                "krw_usd": build_series_item("KRW/USD", "#2563eb", krw_usd_dates, krw_usd_values),
                "chf_usd": build_series_item("CHF/USD", "#16a34a", chf_usd_dates, chf_usd_values),
                "cny_usd": build_series_item("CNY/USD", "#f97316", cny_usd_dates, cny_usd_values),
                "gbp_usd": build_series_item("GBP/USD", "#7c3aed", gbp_usd_dates, gbp_usd_values),
                "eur_usd": build_series_item("EURO/USD", "#0f766e", eur_usd_dates, eur_usd_values),
            },
        },
        "energy": {
            "title": "Crude Oil",
            "subtitle": "WTI combines World Bank monthly spot history from 1982 with daily front-month futures from Yahoo Finance where available. Brent uses daily futures, while Dubai uses World Bank monthly spot.",
            "source": "Yahoo Finance / World Bank Pink Sheet",
            "mode": "raw",
            "connectGaps": True,
            "yAxisLabel": "$ / bbl",
            "formatter": "dollar1",
            "series": {
                "wti": build_series_item("WTI", "#7c3aed", wti_dates, wti_values),
                "brent": build_series_item("Brent", "#2563eb", brent_dates, brent_values),
                "dubai": build_series_item("Dubai", "#f97316", dubai_dates, dubai_values),
            },
        },
        "natural_gas": {
            "title": "Natural Gas / LNG",
            "subtitle": "Henry Hub spot from FRED/EIA and JKM LNG futures close from Yahoo Finance.",
            "source": "FRED / EIA / Yahoo Finance",
            "mode": "raw",
            "yAxisLabel": "$ / MMBtu",
            "formatter": "dollar2",
            "series": {
                "henry_hub": build_series_item("Henry Hub", "#0f766e", henry_hub_dates, henry_hub_values),
                "lng_jkm": build_series_item("JKM LNG", "#2563eb", lng_jkm_dates, lng_jkm_values),
            },
        },
        "metals": {
            "title": "Gold / Silver / Copper",
            "subtitle": "Daily futures closes normalized to 100 at the selected start date for easier cross-asset comparison.",
            "source": "Yahoo Finance",
            "mode": "normalized",
            "yAxisLabel": "Start = 100",
            "formatter": "number1",
            "series": {
                "gold": build_series_item("Gold", "#d97706", gold_dates, gold_values),
                "silver": build_series_item("Silver", "#6b7280", silver_dates, silver_values),
                "copper": build_series_item("Copper", "#b45309", copper_dates, copper_values),
            },
        },
        "strategic": {
            "title": "Strategic Metals",
            "subtitle": "Actual commodity price series normalized to 100. Uranium, nickel, and zinc use public monthly spot/reference series and are carried forward between monthly prints; iron ore uses futures closes. Lithium and tungsten are held out until a stable public spot/futures source is confirmed.",
            "source": "FRED / Yahoo Finance / World Bank Pink Sheet",
            "mode": "normalized",
            "fillMissing": "forward",
            "yAxisLabel": "Start = 100",
            "formatter": "number1",
            "series": {
                "uranium": build_series_item("Uranium U3O8 spot", "#16a34a", uranium_dates, uranium_values),
                "iron_ore": build_series_item("Iron ore 62% Fe (TIO)", "#b45309", iron_ore_dates, iron_ore_values),
                "nickel": build_series_item("Nickel", "#64748b", nickel_dates, nickel_values),
                "zinc": build_series_item("Zinc", "#0ea5e9", zinc_dates, zinc_values),
            },
        },
        "food": {
            "title": "Food Dashboard",
            "subtitle": "2001-01-01 이후 주요 농산물 선물: 옥수수, 대두, 밀(HRW), 밀(SRW), 원당. 원자재별 비교가 쉽도록 선택 시작일을 100으로 정규화합니다.",
            "source": "Yahoo Finance 선물",
            "mode": "normalized",
            "yAxisLabel": "Start = 100",
            "formatter": "number1",
            "series": {
                "corn": build_series_item("옥수수", "#f59e0b", corn_dates, corn_values),
                "soybeans": build_series_item("대두", "#16a34a", soybeans_dates, soybeans_values),
                "wheat_hrw": build_series_item("밀(HRW)", "#dc2626", wheat_hrw_dates, wheat_hrw_values),
                "wheat_srw": build_series_item("밀(SRW)", "#7c3aed", wheat_srw_dates, wheat_srw_values),
                "sugar": build_series_item("원당", "#0f766e", sugar_dates, sugar_values),
            },
        },
    }

    latest_dates = []
    for panel in panels.values():
        for item in panel["series"].values():
            dates = item.get("dates") or []
            if dates:
                latest_dates.append(dates[-1])

    payload = {
        "updatedAt": max(latest_dates) if latest_dates else "",
        "startDate": START_DATE,
        "defaultRange": "3y",
        "ranges": RANGES,
        "longCommodities": {
            "source": "World Bank Pink Sheet monthly prices",
            "series": {
                "wti": build_series_item("WTI", "#16a34a", *long_commodity_series["wti"]),
                "brent": build_series_item("Brent", "#2563eb", *long_commodity_series["brent"]),
                "dubai": build_series_item("Dubai", "#f97316", *long_commodity_series["dubai"]),
                "gold": build_series_item("Gold", "#d97706", *long_commodity_series["gold"]),
                "silver": build_series_item("Silver", "#6b7280", *long_commodity_series["silver"]),
                "copper": build_series_item("Copper", "#b45309", *long_commodity_series["copper"]),
                "iron_ore": build_series_item("Iron Ore", "#b45309", *long_commodity_series["iron_ore"]),
                "nickel": build_series_item("Nickel", "#64748b", *long_commodity_series["nickel"]),
                "zinc": build_series_item("Zinc", "#0ea5e9", *long_commodity_series["zinc"]),
            },
        },
        "panels": panels,
    }

    output_path = Path(__file__).resolve().parents[1] / "data" / "market-macro-data.js"
    output_path.write_text(
        "window.marketMacroData = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
        newline="\n",
    )


if __name__ == "__main__":
    main()
