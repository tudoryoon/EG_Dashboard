from __future__ import annotations

import json
import csv
import re
import tempfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen


DATA_PATH = Path("data/market-macro-data.js")
START_DATE = "1965-01-01"
FOOD_START_DATE = "2001-01-01"
FRED_GRAPH_BASE = "https://fred.stlouisfed.org/graph/fredgraph.csv?id="
FRED_GATEWAY_BASE = "https://www.ivo-welch.info/cgi-bin/fredwrap?symbol="
WORLD_BANK_PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/"
    "related/CMO-Historical-Data-Monthly.xlsx"
)

YAHOO_SERIES = {
    ("energy", "wti"): ("CL=F", START_DATE),
    ("energy", "brent"): ("BZ=F", START_DATE),
    ("natural_gas", "lng_jkm"): ("JKM=F", START_DATE),
    ("metals", "gold"): ("GC=F", START_DATE),
    ("metals", "silver"): ("SI=F", START_DATE),
    ("metals", "copper"): ("HG=F", START_DATE),
    ("strategic", "iron_ore"): ("TIO=F", START_DATE),
    ("food", "corn"): ("ZC=F", FOOD_START_DATE),
    ("food", "soybeans"): ("ZS=F", FOOD_START_DATE),
    ("food", "wheat_hrw"): ("KE=F", FOOD_START_DATE),
    ("food", "wheat_srw"): ("ZW=F", FOOD_START_DATE),
    ("food", "sugar"): ("SB=F", FOOD_START_DATE),
}

FRED_SERIES = {
    ("natural_gas", "henry_hub"): ("DHHNGSP", START_DATE),
    ("strategic", "uranium"): ("PURANUSDM", START_DATE),
}

WORLD_BANK_SERIES = {
    ("energy", "dubai"): ("Crude oil, Dubai", START_DATE),
    ("strategic", "nickel"): ("Nickel", START_DATE),
    ("strategic", "zinc"): ("Zinc", START_DATE),
}


def fetch_text(url: str) -> str:
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    last_error: Exception | None = None
    for _ in range(3):
        try:
            with urlopen(request, timeout=60) as response:  # nosec B310 - fixed public endpoint
                return response.read().decode("utf-8")
        except Exception as error:  # pragma: no cover - network variability
            last_error = error
    raise RuntimeError(f"Failed to fetch {url}") from last_error


def fetch_bytes(url: str) -> bytes:
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    last_error: Exception | None = None
    for _ in range(3):
        try:
            with urlopen(request, timeout=60) as response:  # nosec B310 - fixed public endpoint
                return response.read()
        except Exception as error:  # pragma: no cover - network variability
            last_error = error
    raise RuntimeError(f"Failed to fetch {url}") from last_error


def yahoo_chart_url(symbol: str, start_date: str) -> str:
    period1 = int(datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc).timestamp())
    period2 = int(datetime.now(timezone.utc).timestamp())
    encoded = quote(symbol, safe="")
    return (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{encoded}"
        f"?period1={period1}&period2={period2}&interval=1d&includeAdjustedClose=true&events=div%2Csplits"
    )


def fred_csv_urls(series_id: str, start_date: str) -> list[str]:
    return [
        f"{FRED_GATEWAY_BASE}{series_id}",
        f"{FRED_GRAPH_BASE}{series_id}&cosd={start_date}",
    ]


def parse_yahoo_series(symbol: str, start_date: str) -> tuple[list[str], list[float]]:
    payload = json.loads(fetch_text(yahoo_chart_url(symbol, start_date)))
    result = payload["chart"]["result"][0]
    timestamps = result.get("timestamp") or []
    quote_data = (result.get("indicators") or {}).get("quote") or [{}]
    adjclose_data = (result.get("indicators") or {}).get("adjclose") or [{}]
    closes = adjclose_data[0].get("adjclose") or quote_data[0].get("close") or []

    by_date: dict[str, float] = {}
    for timestamp, close in zip(timestamps, closes):
        if close is None:
            continue
        date_key = (datetime(1970, 1, 1, tzinfo=timezone.utc) + timedelta(seconds=timestamp)).strftime("%Y-%m-%d")
        if date_key < start_date:
            continue
        by_date[date_key] = round(float(close), 4)

    dates = sorted(by_date)
    return dates, [by_date[date] for date in dates]


def parse_fred_series(series_id: str, start_date: str) -> tuple[list[str], list[float]]:
    last_error: Exception | None = None
    for url in fred_csv_urls(series_id, start_date):
        try:
            reader = csv.DictReader(fetch_text(url).splitlines())
            dates: list[str] = []
            values: list[float] = []
            for row in reader:
                lower_row = {str(key).lower(): value for key, value in row.items()}
                raw_date = (row.get("DATE") or lower_row.get("yyyymmdd") or "").strip()
                raw_value = (row.get(series_id) or lower_row.get(series_id.lower()) or "").strip()
                if not raw_date or raw_value in {"", "."}:
                    continue
                if len(raw_date) == 8 and raw_date.isdigit():
                    date_key = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:8]}"
                else:
                    date_key = raw_date[:10]
                if date_key < start_date:
                    continue
                dates.append(date_key)
                values.append(round(float(raw_value), 4))
            if dates:
                return dates, values
        except Exception as error:  # pragma: no cover - network variability
            last_error = error
    if last_error:
        raise last_error
    return [], []


def parse_world_bank_monthly_prices() -> dict[tuple[str, str], tuple[list[str], list[float]]]:
    import openpyxl

    workbook_path = Path(tempfile.gettempdir()) / "CMO-Historical-Data-Monthly.xlsx"
    workbook_path.write_bytes(fetch_bytes(WORLD_BANK_PINK_SHEET_URL))
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook["Monthly Prices"]
    header_row = next(worksheet.iter_rows(min_row=5, max_row=5, values_only=True))
    normalized_headers = [str(value or "").strip().lower() for value in header_row]
    column_index = {
        key: normalized_headers.index(header.lower())
        for key, (header, _start_date) in WORLD_BANK_SERIES.items()
        if header.lower() in normalized_headers
    }
    series = {key: ([], []) for key in WORLD_BANK_SERIES}
    for row in worksheet.iter_rows(min_row=7, values_only=True):
        raw_period = row[0]
        if not raw_period:
            continue
        match = re.match(r"^(\d{4})M(\d{2})$", str(raw_period))
        if not match:
            continue
        date_key = f"{match.group(1)}-{match.group(2)}-01"
        for key, index in column_index.items():
            start_date = WORLD_BANK_SERIES[key][1]
            value = row[index]
            if date_key < start_date or not isinstance(value, (int, float)):
                continue
            series[key][0].append(date_key)
            series[key][1].append(round(float(value), 4))
    try:
        workbook_path.unlink()
    except OSError:
        pass
    return series


def load_payload() -> dict:
    raw = DATA_PATH.read_text(encoding="utf-8").strip()
    prefix = "window.marketMacroData = "
    if not raw.startswith(prefix):
        raise ValueError(f"Unexpected data wrapper in {DATA_PATH}")
    return json.loads(raw[len(prefix) :].rstrip(";"))


def merge_series(existing: dict, dates: list[str], values: list[float], start_date: str) -> dict:
    merged = {
        date: value
        for date, value in zip(existing.get("dates", []), existing.get("values", []))
        if date >= start_date
    }
    merged.update(dict(zip(dates, values)))
    ordered_dates = sorted(merged)
    return {
        **existing,
        "dates": ordered_dates,
        "values": [merged[date] for date in ordered_dates],
    }


def keep_existing_latest(series: dict, label: str, error: Exception) -> str:
    existing_dates = series.get("dates") or []
    if not existing_dates:
        raise RuntimeError(f"No existing commodity observations available for {label}") from error
    latest_date = existing_dates[-1]
    latest_value = (series.get("values") or [""])[-1]
    print(f"WARNING: keeping existing {label} through {latest_date} {latest_value}: {error}")
    return latest_date


def write_payload(payload: dict) -> None:
    text = "window.marketMacroData = "
    text += json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    text += ";\n"
    DATA_PATH.write_text(text, encoding="utf-8")


def main() -> None:
    payload = load_payload()
    latest_dates: list[str] = []

    for (panel_key, series_key), (symbol, start_date) in YAHOO_SERIES.items():
        series = payload["panels"][panel_key]["series"][series_key]
        try:
            dates, values = parse_yahoo_series(symbol, start_date)
            if not dates:
                raise RuntimeError(f"No Yahoo observations returned for {panel_key}.{series_key} ({symbol})")
        except Exception as error:  # pragma: no cover - network variability
            latest_dates.append(keep_existing_latest(series, f"{panel_key}.{series_key} ({symbol})", error))
            continue
        payload["panels"][panel_key]["series"][series_key] = merge_series(series, dates, values, start_date)
        updated_series = payload["panels"][panel_key]["series"][series_key]
        latest_dates.append(updated_series["dates"][-1])
        print(f"{panel_key}.{series_key}: {updated_series['dates'][-1]} {updated_series['values'][-1]}")

    for (panel_key, series_key), (series_id, start_date) in FRED_SERIES.items():
        series = payload["panels"][panel_key]["series"][series_key]
        try:
            dates, values = parse_fred_series(series_id, start_date)
            if not dates:
                raise RuntimeError(f"No FRED observations returned for {panel_key}.{series_key} ({series_id})")
        except Exception as error:  # pragma: no cover - network variability
            latest_dates.append(keep_existing_latest(series, f"{panel_key}.{series_key} ({series_id})", error))
            continue
        payload["panels"][panel_key]["series"][series_key] = merge_series(series, dates, values, start_date)
        updated_series = payload["panels"][panel_key]["series"][series_key]
        latest_dates.append(updated_series["dates"][-1])
        print(f"{panel_key}.{series_key}: {updated_series['dates'][-1]} {updated_series['values'][-1]}")

    try:
        world_bank_series = parse_world_bank_monthly_prices()
    except Exception as error:  # pragma: no cover - network variability
        print(f"WARNING: World Bank Pink Sheet fetch failed; preserving existing monthly series: {error}")
        world_bank_series = {key: ([], []) for key in WORLD_BANK_SERIES}
    for (panel_key, series_key), (_header, start_date) in WORLD_BANK_SERIES.items():
        series = payload["panels"][panel_key]["series"][series_key]
        try:
            dates, values = world_bank_series[(panel_key, series_key)]
            if not dates:
                raise RuntimeError(f"No World Bank observations returned for {panel_key}.{series_key}")
        except Exception as error:  # pragma: no cover - network variability
            latest_dates.append(keep_existing_latest(series, f"{panel_key}.{series_key}", error))
            continue
        payload["panels"][panel_key]["series"][series_key] = merge_series(series, dates, values, start_date)
        updated_series = payload["panels"][panel_key]["series"][series_key]
        latest_dates.append(updated_series["dates"][-1])
        print(f"{panel_key}.{series_key}: {updated_series['dates'][-1]} {updated_series['values'][-1]}")

    payload["panels"]["energy"]["fillMissing"] = "forward"
    payload["panels"]["energy"]["subtitle"] = (
        "WTI combines World Bank monthly spot history from 1982 with daily front-month futures from Yahoo Finance. "
        "Brent uses daily futures; Dubai uses World Bank monthly spot and is forward-filled between monthly prints."
    )
    payload["panels"]["natural_gas"]["fillMissing"] = "forward"
    payload["panels"]["natural_gas"]["subtitle"] = (
        "Henry Hub spot from FRED/EIA and JKM LNG futures close from Yahoo Finance, forward-filled across missing days."
    )

    payload["updatedAt"] = max([payload.get("updatedAt", "")] + latest_dates)
    write_payload(payload)
    print(f"Updated commodity dashboards through {max(latest_dates)}")


if __name__ == "__main__":
    main()
